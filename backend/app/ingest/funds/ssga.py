"""State Street (SSGA) issuer holdings adapter for XLV / XLF / XBI.

SSGA publishes full daily holdings as public XLSX files
(``holdings-daily-us-en-<symbol>.xlsx``). The sheet has a preamble (fund name,
ticker, an "As of DD-Mon-YYYY" line, sometimes net assets / expense ratio)
before the actual header row, so the parser scans for the header instead of
assuming a fixed layout.

CAVEAT: finance hosts are blocked at this sandbox's egress proxy, so the
layout handled below follows SSGA's documented file shape and is parsed
defensively. Adapters never raise: failures log a warning and return None.
"""
import re
from datetime import date, datetime
from io import BytesIO
from zipfile import BadZipFile

import httpx
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.logging import get_logger
from app.ingest.funds.base import HoldingRow, HoldingsSnapshot

logger = get_logger(__name__)
REQUEST_TIMEOUT = httpx.Timeout(30.0)
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}

PREAMBLE_SCAN_ROWS = 15
AS_OF_DATE_FORMATS = (
    "%d-%b-%Y",  # 31-Dec-2025
    "%d-%B-%Y",  # 31-December-2025
    "%B %d %Y",  # December 31 2025
    "%B %d, %Y",  # December 31, 2025
    "%m/%d/%Y",
    "%Y-%m-%d",
)
PARSE_ERRORS = (
    KeyError,
    IndexError,
    TypeError,
    ValueError,
    AttributeError,
    BadZipFile,
    InvalidFileException,
)


async def fetch_ssga_holdings(base_url: str, symbol: str) -> HoldingsSnapshot | None:
    url = f"{base_url}/holdings-daily-us-en-{symbol.lower()}.xlsx"
    try:
        async with httpx.AsyncClient(
            timeout=REQUEST_TIMEOUT, headers=HEADERS, follow_redirects=True
        ) as client:
            response = await client.get(url)
            response.raise_for_status()
    except httpx.HTTPError as exc:
        logger.warning("ssga_fetch_failed", symbol=symbol, url=url, error=str(exc))
        return None

    snapshot = parse_ssga_xlsx(response.content, source_url=url)
    if snapshot is None:
        logger.warning("ssga_parse_failed", symbol=symbol, url=url)
    else:
        logger.info(
            "ssga_fetch_ok",
            symbol=symbol,
            rows=len(snapshot.rows),
            as_of=str(snapshot.as_of),
        )
    return snapshot


def _cell_text(value) -> str:
    return value.strip() if isinstance(value, str) else ""


def _to_float(value) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace("%", "").replace(",", "").strip())
        except ValueError:
            return None
    return None


def _parse_as_of_text(text: str) -> date | None:
    """Extract a date from an 'As of ...' preamble cell."""
    match = re.search(r"as\s+of[:\s]*(.+)", text, flags=re.IGNORECASE)
    candidate = (match.group(1) if match else text).strip().rstrip(".")
    for fmt in AS_OF_DATE_FORMATS:
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    return None


def _find_as_of(rows: list[tuple]) -> date | None:
    for row in rows[:PREAMBLE_SCAN_ROWS]:
        if not any("as of" in _cell_text(cell).lower() for cell in row):
            continue
        # A datetime cell in the same row beats text parsing.
        for cell in row:
            if isinstance(cell, datetime):
                return cell.date()
            if isinstance(cell, date):
                return cell
        for cell in row:
            parsed = _parse_as_of_text(_cell_text(cell))
            if parsed is not None:
                return parsed
    return None


def _find_header(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """Locate the header row (has 'name' plus 'weight' or 'ticker') and map
    the column indexes of interest."""
    for row_idx, row in enumerate(rows[: PREAMBLE_SCAN_ROWS + 5]):
        cells = [_cell_text(cell).lower() for cell in row]
        has_name = any("name" in cell for cell in cells)
        has_weight_or_ticker = any("weight" in cell or "ticker" in cell for cell in cells)
        if not (has_name and has_weight_or_ticker):
            continue
        columns: dict[str, int] = {}
        for col_idx, cell in enumerate(cells):
            if "name" in cell and "name" not in columns:
                columns["name"] = col_idx
            elif "ticker" in cell and "ticker" not in columns:
                columns["ticker"] = col_idx
            elif "weight" in cell and "weight" not in columns:
                columns["weight"] = col_idx
            elif "sector" in cell and "sector" not in columns:
                columns["sector"] = col_idx
        if "name" in columns and "weight" in columns:
            return row_idx, columns
    return None


def _get(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_ssga_xlsx(content: bytes, source_url: str = "") -> HoldingsSnapshot | None:
    """Pure defensive parse of an SSGA daily-holdings XLSX."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]

        as_of = _find_as_of(rows)
        header = _find_header(rows)
        if header is None:
            logger.warning("ssga_header_not_found")
            return None
        header_idx, columns = header

        parsed: list[tuple[str, str | None, float, str | None]] = []
        for row in rows[header_idx + 1 :]:
            if all(cell is None or _cell_text(cell) == "" for cell in row):
                break  # first fully-empty row ends the data block
            name = _cell_text(_get(row, columns.get("name")))
            weight = _to_float(_get(row, columns.get("weight")))
            if not name or weight is None:
                break  # footer text (disclaimers etc.)
            ticker = _cell_text(_get(row, columns.get("ticker"))) or None
            sector = _cell_text(_get(row, columns.get("sector"))) or None
            parsed.append((name, ticker, weight, sector))
    except PARSE_ERRORS as exc:
        logger.warning("ssga_xlsx_parse_failed", error=str(exc))
        return None

    if not parsed:
        return None

    # Some exports carry weights as fractions (sum ~= 1.0) instead of percents.
    total = sum(weight for _, _, weight, _ in parsed)
    if 0 < total <= 1.5:
        parsed = [(name, ticker, weight * 100.0, sector) for name, ticker, weight, sector in parsed]

    if as_of is None:
        # Do NOT invent a date silently -- log so the service can label it.
        logger.warning("ssga_as_of_missing")
        as_of = date.today()

    parsed.sort(key=lambda item: item[2], reverse=True)
    holding_rows = [
        HoldingRow(rank=rank, holding_name=name, weight_pct=weight, ticker=ticker, sector=sector)
        for rank, (name, ticker, weight, sector) in enumerate(parsed, start=1)
    ]
    return HoldingsSnapshot(
        as_of=as_of,
        source="ssga",
        source_url=source_url,
        rows=holding_rows,
        is_top10_only=False,
    )


def parse_ssga_fund_facts(content: bytes) -> dict:
    """Best-effort fund-level extras from the preamble/summary rows.

    Returns {"aum": float|None, "expense_ratio": float|None}; both None when
    the sheet carries no such rows (or cannot be read at all).
    """
    facts: dict = {"aum": None, "expense_ratio": None}
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        for row_num, row in enumerate(workbook.active.iter_rows(values_only=True)):
            if row_num >= PREAMBLE_SCAN_ROWS:
                break
            label = " ".join(_cell_text(cell).lower() for cell in row if _cell_text(cell))
            numeric = next(
                (v for v in (_to_float(cell) for cell in row) if v is not None), None
            )
            if numeric is None:
                continue
            if facts["aum"] is None and ("net assets" in label or "aum" in label):
                facts["aum"] = numeric
            elif facts["expense_ratio"] is None and "expense ratio" in label:
                facts["expense_ratio"] = numeric
    except PARSE_ERRORS as exc:
        logger.warning("ssga_fund_facts_parse_failed", error=str(exc))
    return facts
